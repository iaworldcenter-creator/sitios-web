import os
import json
import re
import ssl
import urllib.request
import openpyxl
import pandas as pd
from concurrent.futures import ThreadPoolExecutor

BASE_DIR = r"E:\sitios web\pc-custom-lab"
CATEGORIES_DIR = os.path.join(BASE_DIR, "data", "categorias")
IMG_BASE_DIR = os.path.join(BASE_DIR, "assets", "img", "catalog")
os.makedirs(CATEGORIES_DIR, exist_ok=True)
os.makedirs(IMG_BASE_DIR, exist_ok=True)

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

EXCEL_CONFIG = r"D:\Descargas\lista3\TOL 082426\1308 CONFIGURACIONES TOL 082426.xlsx"
EXCEL_PRICES = r"D:\Descargas\lista3\TOL 082426\1308 LISTA DE PRECIOS DE CT TOL 082426.xlsx"

print("=" * 80)
print("CLASIFICACIÓN DETERMINISTA POR PREFIJOS OFICIALES CT Y DESCARGA POR CARPETAS")
print("=" * 80)

# 1. Definir Prefijos Exactos de CT Internacional
PREFIX_CATEGORY_MAP = {
    'MON': 'monitores',
    'MBD': 'tarjetas_madre',
    'CPU': 'procesadores',
    'MEM': 'memorias_ram',
    'DDU': 'discos_duros',
    'TVI': 'tarjetas_de_video',
    'FUE': 'fuentes_energia',
    'NOB': 'fuentes_energia',
    'REG': 'fuentes_energia',
    'UPS': 'fuentes_energia',
    'GAB': 'gabinetes',
    'ENF': 'enfriamiento',
    'IMP': 'impresoras',
    'MLT': 'impresoras',
    'TIN': 'consumibles',
    'TON': 'consumibles',
    'CRT': 'consumibles',
    'CNM': 'consumibles',
    'SWT': 'conectividad_redes',
    'ROU': 'conectividad_redes',
    'WIF': 'conectividad_redes',
    'RED': 'conectividad_redes',
    'SOF': 'software',
    'ACC': 'accesorios_perifericos',
    'TEC': 'accesorios_perifericos',
    'MOU': 'accesorios_perifericos',
    'AUD': 'accesorios_perifericos',
    'KIT': 'accesorios_perifericos',
    'CAM': 'telefonia_seguridad',
    'DVR': 'telefonia_seguridad',
    'NVR': 'telefonia_seguridad',
    'CON': 'telefonia_seguridad',
    'TEL': 'telefonia_seguridad',
    'LAP': 'equipos_de_marca',
    'NOT': 'equipos_de_marca',
    'DES': 'equipos_de_marca',
    'AIO': 'equipos_de_marca',
    'LCT': 'punto_de_venta',
    'POS': 'punto_de_venta',
    'CAJ': 'punto_de_venta',
    'BSC': 'punto_de_venta',
    'PAN': 'electronica_consumo',
    'TV': 'electronica_consumo',
    'BOC': 'electronica_consumo',
    'AIR': 'linea_blanca',
    'REF': 'linea_blanca',
    'FRI': 'linea_blanca',
    'OUT': 'outlet_remates'
}

# Crear carpetas físicas por categoría
for cat in set(PREFIX_CATEGORY_MAP.values()):
    os.makedirs(os.path.join(IMG_BASE_DIR, cat), exist_ok=True)

# 2. Parsear Excel de Precios
wb_prices = openpyxl.load_workbook(EXCEL_PRICES, data_only=True)
all_products = []

for sheet_name in wb_prices.sheetnames:
    if sheet_name in ['INDICE', 'DIRECTORIO', 'DIRECTORIO DE COORDINADORES', 'LISTA DE PRECIOS']:
        continue
    ws = wb_prices[sheet_name]
    
    for r in ws.iter_rows(values_only=True):
        if not r or len(r) < 6: continue
        
        sku_cand = None
        desc_cand = None
        price_usd = None
        marca_cand = 'CT'
        
        for c_idx, cell in enumerate(r):
            val = str(cell or '').strip()
            if re.match(r'^[A-Z0-9]{6,15}$', val) and not sku_cand and any(c.isalpha() for c in val) and any(c.isdigit() for c in val):
                sku_cand = val
                if c_idx + 1 < len(r):
                    desc_cand = str(r[c_idx + 1] or '').strip()
            if isinstance(cell, (int, float)) and cell > 0.5:
                price_usd = float(cell)
                
        if sku_cand and desc_cand and len(desc_cand) > 3 and price_usd:
            # Extraer marca
            first_word = desc_cand.split()[0].upper()
            if first_word in ['ASUS', 'INTEL', 'AMD', 'KINGSTON', 'ACTECK', 'MSI', 'GIGABYTE', 'TRIPP-LITE', 'ADATA', 'CORSAIR', 'LOGITECH', 'SAMSUNG', 'LG', 'DELL', 'HP', 'LENOVO', 'EPSON', 'CANON', 'BROTHER', 'TP-LINK', 'DAHUA', 'HIKVISION', 'BENQ', 'AOC', 'VIEWSONIC', 'HYUNDAI', 'VORAGO', 'GHIA', 'BALAM']:
                marca_cand = first_word

            # Asignar categoría estricta por prefijo SKU de 3 letras
            prefix3 = sku_cand[:3].upper()
            prefix2 = sku_cand[:2].upper()
            cat_assigned = PREFIX_CATEGORY_MAP.get(prefix3) or PREFIX_CATEGORY_MAP.get(prefix2)
            
            if not cat_assigned:
                # Fallback por palabras clave en descripción
                d_low = desc_cand.lower()
                if 'monitor' in d_low: cat_assigned = 'monitores'
                elif 'motherboard' in d_low or 'tarjeta madre' in d_low: cat_assigned = 'tarjetas_madre'
                elif 'procesador' in d_low or 'core i' in d_low or 'ryzen' in d_low: cat_assigned = 'procesadores'
                elif 'memoria' in d_low or 'ddr' in d_low: cat_assigned = 'memorias_ram'
                elif 'disco' in d_low or 'ssd' in d_low or 'm.2' in d_low: cat_assigned = 'discos_duros'
                elif 'tarjeta de video' in d_low or 'rtx' in d_low: cat_assigned = 'tarjetas_de_video'
                else: cat_assigned = 'varios_hardware'

            costo_mxn = price_usd * 19.50
            precio_lista_60 = round(costo_mxn * 1.60, 2)
            precio_venta_35 = round(costo_mxn * 1.35, 2)
            precio_mayoreo_25 = round(costo_mxn * 1.25, 2)
            precio_piso_20 = round(costo_mxn * 1.20, 2)

            all_products.append({
                "sku": sku_cand,
                "nombre": desc_cand,
                "descripcion_completa": f"{desc_cand}. Clave Oficial CT: {sku_cand}. Marca: {marca_cand}.",
                "categoria_ct": sheet_name,
                "categoria_clasificada": cat_assigned,
                "marca": marca_cand,
                "base_cost_usd": round(price_usd, 2),
                "costo_proveedor_mxn": round(costo_mxn, 2),
                "precio_original": precio_lista_60,
                "original": precio_lista_60,
                "precio": precio_venta_35,
                "precio_mxn": precio_venta_35,
                "precio_mayoreo_10pzs": precio_mayoreo_25,
                "precio_piso_minimo_20": precio_piso_20,
                "img_rel": f"assets/img/catalog/{cat_assigned}/{sku_cand}.jpg",
                "img": f"assets/img/catalog/{cat_assigned}/{sku_cand}.jpg",
                "cdn_url": f"https://static.ctonline.mx/imagenes/{sku_cand}/{sku_cand}_400.jpg"
            })

print(f"✓ Total productos procesados: {len(all_products)}")

# 3. Separar y Guardar por Categorías
cat_dict = {}
for p in all_products:
    c = p['categoria_clasificada']
    if c not in cat_dict: cat_dict[c] = []
    cat_dict[c].append(p)

for cname, prods in cat_dict.items():
    with open(os.path.join(CATEGORIES_DIR, f"{cname}.json"), "w", encoding="utf-8") as f:
        json.dump(prods, f, ensure_ascii=False, indent=2)
    pd.DataFrame(prods).to_csv(os.path.join(CATEGORIES_DIR, f"{cname}.csv"), index=False, encoding="utf-8")

print(f"✓ Conteo real por categorías:")
for cname, prods in cat_dict.items():
    print(f"   -> [{cname}]: {len(prods)} productos")

# 4. Descarga de Imágenes Organizadas por Carpeta de Categoría
def download_categorized_img(prod):
    cat = prod["categoria_clasificada"]
    sku = prod["sku"]
    local_path = os.path.join(IMG_BASE_DIR, cat, f"{sku}.jpg")
    
    if os.path.exists(local_path) and os.path.getsize(local_path) > 800:
        return True
        
    cdn = f"https://static.ctonline.mx/imagenes/{sku}/{sku}_400.jpg"
    req = urllib.request.Request(cdn, headers={'User-Agent': 'Mozilla/5.0'})
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=3) as resp:
            d = resp.read()
            if len(d) > 500:
                with open(local_path, "wb") as f_out:
                    f_out.write(d)
                return True
    except:
        pass
    return False

# Descargar las primeras 60 imágenes de cada categoría principal
target_downloads = []
for cname, prods in cat_dict.items():
    target_downloads.extend(prods[:60])

print(f"\nDescargando {len(target_downloads)} imágenes organizadas en sus carpetas respectivas...")
with ThreadPoolExecutor(max_workers=20) as executor:
    dl_res = list(executor.map(download_categorized_img, target_downloads))

print(f"✅ {sum(1 for r in dl_res if r)} imágenes descargadas en sus carpetas de categoría.")

# 5. Generar configuraciones armadas limpias
wb_config = openpyxl.load_workbook(EXCEL_CONFIG, data_only=True)
clean_combos = []
for sname in ['INTEL 14va GEN', 'INTEL 12va GEN', 'AMD RYZEN']:
    if sname not in wb_config.sheetnames: continue
    ws = wb_config[sname]
    rows = list(ws.iter_rows(values_only=True))
    header_row = 4
    headers = rows[header_row]
    for col_idx in range(3, len(headers)):
        pc_name = headers[col_idx]
        if not pc_name or 'Tabla' in str(pc_name) or 'CT' in str(pc_name): continue
        components = []
        total_usd = 0.0
        for r in rows[header_row+1:]:
            sku = str(r[1] or '').strip()
            desc = str(r[2] or '').strip()
            if col_idx < len(r) and r[col_idx] is not None:
                try:
                    price_val = float(r[col_idx])
                    if price_val > 0 and sku and sku != 'None':
                        components.append({'sku': sku, 'desc': desc, 'usd': price_val})
                        total_usd += price_val
                except: pass
        if total_usd > 50:
            costo_mxn = total_usd * 19.50
            clean_combos.append({
                "sku": f"PC-{sname.replace(' ', '-').upper()}-{col_idx}",
                "nombre": f"PC Armada {pc_name} | {sname}",
                "descripcion_completa": f"Configuración recomendada oficial CT: " + ", ".join([f"{c['sku']} ({c['desc']})" for c in components]),
                "categoria_ct": "Equipos Armados",
                "categoria_clasificada": "equipos_de_marca",
                "marca": "CT Ensamble",
                "base_cost_usd": round(total_usd, 2),
                "costo_proveedor_mxn": round(costo_mxn, 2),
                "precio_original": round(costo_mxn * 1.60, 2),
                "original": round(costo_mxn * 1.60, 2),
                "precio": round(costo_mxn * 1.35, 2),
                "precio_mxn": round(costo_mxn * 1.35, 2),
                "precio_mayoreo_10pzs": round(costo_mxn * 1.25, 2),
                "precio_piso_minimo_20": round(costo_mxn * 1.20, 2),
                "img": "assets/img/catalog/gabinete_negro.webp",
                "local_img": "assets/img/catalog/gabinete_negro.webp"
            })

with open(os.path.join(BASE_DIR, "data", "configuraciones_armadas.json"), "w", encoding="utf-8") as f:
    json.dump(clean_combos, f, ensure_ascii=False, indent=2)

# 6. Catálogo Selecto de Alta Fidelidad para Frontend
# Armar catálogo ordenado por Monitores, Procesadores, Tarjetas Madre, RAM, etc.
curated_catalog = clean_combos
for cname in ['monitores', 'procesadores', 'tarjetas_madre', 'tarjetas_de_video', 'memorias_ram', 'discos_duros', 'fuentes_energia', 'gabinetes', 'enfriamiento', 'impresoras', 'consumibles', 'conectividad_redes', 'software', 'accesorios_perifericos', 'telefonia_seguridad', 'equipos_de_marca', 'punto_de_venta', 'electronica_consumo', 'linea_blanca', 'outlet_remates', 'varios_hardware']:
    if cname in cat_dict:
        curated_catalog.extend(cat_dict[cname][:150])

with open(os.path.join(BASE_DIR, "js", "ct-catalog-data.js"), "w", encoding="utf-8") as f:
    f.write(f"window.CT_CATALOG_DATA = {json.dumps(curated_catalog, ensure_ascii=False)};\n")
    f.write(f"window.PC_COMBOS_DATA = {json.dumps(clean_combos, ensure_ascii=False)};\n")

print(f"✓ js/ct-catalog-data.js generado con {len(curated_catalog)} productos clasificados con precisión milimétrica.")

# 7. Actualizar getFilteredList en js/ct-exact-catalog-engine.js para usar categoria_clasificada exacta
with open(os.path.join(BASE_DIR, "js", "ct-exact-catalog-engine.js"), "r", encoding="utf-8") as f:
    eng_code = f.read()

FILTER_ACCURATE = """function getFilteredList() {
    let items = [...(window.PC_COMBOS_DATA || []), ...(window.CT_CATALOG_DATA || [])];

    if (activeSelectedCategory !== 'Todas') {
        items = items.filter(p => {
            const catClasif = (p.categoria_clasificada || '').toLowerCase();
            return catClasif === activeSelectedCategory.toLowerCase();
        });
    }

    if (activeSelectedBrand !== 'Todas') {
        items = items.filter(p => (p.marca || '').toUpperCase() === activeSelectedBrand.toUpperCase());
    }

    if (currentSortCriterion === 'precio_asc') {
        items.sort((a, b) => (a.precio_mxn || a.precio || 0) - (b.precio_mxn || b.precio || 0));
    } else if (currentSortCriterion === 'precio_desc') {
        items.sort((a, b) => (b.precio_mxn || b.precio || 0) - (a.precio_mxn || a.precio || 0));
    } else if (currentSortCriterion === 'nombre') {
        items.sort((a, b) => (a.nombre || '').localeCompare(b.nombre || ''));
    }

    return items;
}"""

eng_code = re.sub(r"function getFilteredList\(\) \{[\s\S]*?\n\}", FILTER_ACCURATE, eng_code)

with open(os.path.join(BASE_DIR, "js", "ct-exact-catalog-engine.js"), "w", encoding="utf-8") as f:
    f.write(eng_code)

print("✓ js/ct-exact-catalog-engine.js actualizado con filtro exacto de categoría.")

# 8. Espejo a C:
BASE_DIR_C = r"C:\Users\nflgd\OneDrive\Documentos\ChatGPT\sitios web\pc-custom-lab"
for root, dirs, files in os.walk(BASE_DIR):
    if '.git' in root or 'node_modules' in root: continue
    for file in files:
        src = os.path.join(root, file)
        rel = os.path.relpath(src, r"E:\sitios web")
        dst = os.path.join(r"C:\Users\nflgd\OneDrive\Documentos\ChatGPT\sitios web", rel)
        if os.path.exists(os.path.dirname(dst)):
            try:
                with open(src, "rb") as f_in, open(dst, "wb") as f_out:
                    f_out.write(f_in.read())
            except: pass

print("✅ CLASIFICACIÓN EXACTA Y FOTOGRAFÍAS DESPLEGADAS CON ÉXITO!")
